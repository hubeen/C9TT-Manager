'''
    made by mhubeen (문승현)

    공주대학교 클라우드 나인 시간표 만들어주는 프로젝트

'''

import openpyxl as mh
import json
name_row = "C"
d_offset_s = "F4"
d_offset_l = "O4"
n_offset = "C4"
alpha = ['E','F', 'G','H','I','J','K','L','M','N','O','P','Q','R','S','T']

one_sh = [] #시스트에 대한 리스트
two_sh = [] #시스트에 대한 리스트
all_sh = [one_sh, two_sh] #전체 시트에 대한 내용
one_dy = {} # 1호점 사람들 정보
two_dy = {} # 2호점 사람들 정보
all_dy = [one_dy, two_dy] #전체 사람들에 대한 내용
all_names = []
max_row = 0
max_col = 0
max_day = 0
one_td = 0
two_td = 0
all_td = [one_td, two_td]

p_cnt = 0
def get_row(i, max):
    # 요일 개수를 통해 요일마다 인원 수에 맞게 시작하는 위치를 반환해주는  함수
    # argv 1 : 요일 갯수
    # argv 2 : 최대 인원 수
    if(0 == i):
        return 4
    else:
        return max+(get_row(i-1, max)+1)

def get_col(i):
    # 1호점과 2호점의 마감 시간이 다르기 때문에 시간을 반환해주는 함수
    # argv 1 : 호점
    global all_sh
    global wb
    ws = wb[all_sh[i][0]]
    t = ws[3] # 시간 적힌 공간
    cnt = -1
    sw = False
    for i in t:
        tmp_s = i.value
        if(tmp_s == '8~9'):
            sw = True

        if(sw == True):
            cnt += 1
            if(tmp_s == '계'):
                sw = False

    return cnt

def get_day(ho_i, sh_i):
    global wb
    global max_row
    # 시트에 요일 갯수 반환해주는 함수
    ws = wb[all_sh[ho_i][sh_i]]
    max_row = ws.max_row
    ret = 0
    day_s = ws["A1":"A" + str(max_row)]
    for co in day_s:
        for ce in co:
            tmp_s = ce.value
            if(tmp_s is not None):
                if(tmp_s in "계"):
                    ret += 1
    return ret

def get_sheet():
    global one_sh
    global two_sh
    # 1호점 시트와 2호점 시트를 구분하여 배열에 담는 함수
    for n in sl:
        if("1호점" in n):
            one_sh.append(n)
        if("2호점" in n):
            two_sh.append(n)

def get_people(ar_day):
    global p_cnt
    global wb
    global max_row
    global all_sh
    global one_sh
    global two_sh
    global n_offset
    global all_dy
    global one_dy
    global two_dy
    global all_names
    sw = True
    ws = wb[one_sh[0]]
    names = ws[n_offset:"C" + str(max_row)]
    for co in names:
        for ce in co:
            tmp_s = ce.value
            if(tmp_s is not None):
                p_cnt += 1
                all_names.append(tmp_s)
            else:
                sw = False
                break
        if(sw == False):
            break
    for dy in all_dy:
        for i in range(len(all_names)):
            days = []
            dy[all_names[i]] = days

def sh_parsing(ho_i, sh_i):
    #시트 긁어오는 함수
    #argv 1 : 1호점 시트인지 2호점 시트인지 확인하는 값
    #argv 2 : 몇 호점의 시트를 가져올 건지
    global wb
    global all_dy
    global one_dy
    global two_dy
    global all_sh
    global one_sh
    global two_sh
    global all_names
    global max_day
    global one_td
    global two_td
    global all_td
    global alpha

    all_td[ho_i] = get_col(ho_i)
    len_n = len(all_names)
    #len_s = len(all_)
    for n in range(max_day):
        tmp_r = get_row(n, len_n) # return int type
        for c in range(len(all_names)):
            tmp_d = wb[all_sh[ho_i][sh_i]]['F' + str(tmp_r+c) : alpha[all_td[ho_i]] + str(tmp_r+c)]
            tmp_v = []
            for co in tmp_d:
                for ce in co:
                    tmp_s = ce.value
                    if(tmp_s is None):
                        tmp_v.append("0")
                    else:
                        tmp_v.append(tmp_s)
            all_dy[ho_i][all_names[c]].append(tmp_v)

def dy_parsing(ho_i):
    global all_dy
    global all_names
    global all_td
    tt_arr = {}

    for i in range(6): # 요일 [[][][][][][][][][][][][][][]],[₩~~~]
        tmp_d = [[] for i in range(all_td[ho_i])]
        tt_arr[i] = tmp_d

    for yoil in range(len(all_dy[ho_i]['이채연'])): #월~토
        for arr_i in range(len(tt_arr[yoil])): # [[], [], [], [], [], [], [], [], [], []]
            for name in all_names:
                hour = all_dy[ho_i][name][yoil][arr_i]
                if(hour != '0'):
                    tt_arr[yoil][arr_i].append(name + "(" + str(hour) + ")")

    return tt_arr

if __name__ == "__main__":
    wb = mh.load_workbook("./sample/example.xlsx")
    sl = wb.sheetnames
    get_sheet()

    max_day = get_day(0,0)
    get_people(max_day)
    sh_parsing(0, 0)
    sh_parsing(1, 0)

    with open("test.json", "w", encoding='UTF-8-sig') as json_file:
        json_file.write(json.dumps(dy_parsing(0), ensure_ascii=False))

    #dy_parsing(1)

wb.close()
